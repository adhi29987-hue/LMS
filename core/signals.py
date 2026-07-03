import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.db.models.signals import post_save, post_delete
from django.dispatch import receiver
from django.conf import settings
from .models import Book, Student, Issue, Attendance


# Use /tmp in production (read-only filesystem) or BASE_DIR locally
if os.environ.get('DATABASE_URL'):
    EXCEL_PATH = os.path.join('/tmp', 'LMS_Data.xlsx')
else:
    EXCEL_PATH = os.path.join(settings.BASE_DIR, 'LMS_Data.xlsx')


def get_border():
    thin = Side(style='thin', color='D1D5DB')
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def style_header(ws, headers, fill_color='1A56DB'):
    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
    font = Font(bold=True, color='FFFFFF', size=11)
    border = get_border()
    ws.append(headers)
    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    ws.row_dimensions[1].height = 22


def style_data_row(ws, row_num, num_cols):
    border = get_border()
    fill = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid') if row_num % 2 == 0 else PatternFill(fill_type=None)
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.border = border
        cell.fill = fill
        cell.alignment = Alignment(vertical='center')


def auto_fit_columns(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 4, 40)


def generate_full_excel():
    wb = openpyxl.Workbook()

    # ── Books Sheet ──
    ws_books = wb.active
    ws_books.title = 'Books'
    style_header(ws_books, ['Book ID', 'Title', 'Author', 'Status', 'Issued To (ID)', 'Issued To (Name)', 'Due Date'])
    for i, book in enumerate(Book.objects.all(), 2):
        student_id = '-'
        student_name = '-'
        due_date = '-'
        if book.status == 'Issued':
            active_issue = Issue.objects.filter(book=book, is_returned=False).order_by('-issue_date').first()
            if active_issue:
                student_id = active_issue.student.student_id
                student_name = active_issue.student.name
                due_date = str(active_issue.return_date)
        
        ws_books.append([book.book_id, book.title, book.author, book.status, student_id, student_name, due_date])
        style_data_row(ws_books, i, 7)
    auto_fit_columns(ws_books)

    # ── Students Sheet ──
    ws_students = wb.create_sheet('Students')
    style_header(ws_students, ['Student ID', 'Name', 'Department'], fill_color='10B981')
    for i, s in enumerate(Student.objects.all(), 2):
        ws_students.append([s.student_id, s.name, s.department])
        style_data_row(ws_students, i, 3)
    auto_fit_columns(ws_students)

    # ── Issues Sheet ──
    ws_issues = wb.create_sheet('Issues')
    style_header(ws_issues, ['Issue ID', 'Student', 'Book', 'Issue Date', 'Return Date', 'Status'], fill_color='F59E0B')
    for i, iss in enumerate(Issue.objects.select_related('student', 'book').all(), 2):
        status = 'Returned' if iss.is_returned else 'Issued'
        ws_issues.append([
            iss.pk,
            iss.student.name,
            iss.book.title,
            str(iss.issue_date),
            str(iss.return_date),
            status
        ])
        style_data_row(ws_issues, i, 6)
    auto_fit_columns(ws_issues)

    # ── Attendance Sheet ──
    ws_att = wb.create_sheet('Attendance')
    style_header(ws_att, ['ID', 'Name', 'Role', 'Time In', 'Status'], fill_color='6366F1')
    for i, att in enumerate(Attendance.objects.all(), 2):
        ws_att.append([att.pk, att.name, att.role, str(att.time_in), att.status])
        style_data_row(ws_att, i, 5)
    auto_fit_columns(ws_att)

    wb.save(EXCEL_PATH)


# Django Signals — auto-regenerate Excel on any model change
@receiver([post_save, post_delete], sender=Book)
def on_book_change(sender, **kwargs):
    generate_full_excel()


@receiver([post_save, post_delete], sender=Student)
def on_student_change(sender, **kwargs):
    generate_full_excel()


@receiver([post_save, post_delete], sender=Issue)
def on_issue_change(sender, **kwargs):
    generate_full_excel()


@receiver([post_save, post_delete], sender=Attendance)
def on_attendance_change(sender, **kwargs):
    generate_full_excel()
